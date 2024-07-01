import openpyxl
from openpyxl.utils import get_column_letter
import json
from sys import argv
from math import inf
import pdb

# Helper class to represent a table in the spreadsheet
class Table:
    def __init__(self):
        self.column_headers = []
        self.row_names = []
        self.last_row = []
        self.corner_coords = []


# Helper class to represent a cell in the spreadsheet
class Cell:
    def __init__(self, type, value, row, col):
        self.type = type # "None", "Column", "Row Name", "Last Row"
        self.value = value # self.ws.cell(row=self.row, column=self.col).value
        self.row = row
        self.col = col
    def __repr__(self):
        return f"{self.type}: {self.value}"

# Main Excel Processing Class
class ExcelProcessor:
    def __init__(self, xlsx_file, json_file):
        self.wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        self.ws = self.wb.active
        self.json_file = json_file
        self.max_row = self.ws.max_row
        self.max_col = self.ws.max_column
        self.matrix =  [[None for _ in range(self.max_col)] for _ in range(self.max_row)]
        self.json_data = None
        self.global_col_coordinates = []
        self.tables = []
        # for all such coordinates found
        self.columns_found = []
        self.row_names_found = []
        self.last_row_found = []
        self.corner_coordinates = []


    def print_workbook_contents(self):
        # Iterate through each row in the worksheet
        for row in self.ws.iter_rows(values_only=True):
            # Print each cell value in the row
            print(row)
        
    def get_row(self,row):
        """ Converts 1-indexed row to 0-indexed row."""
        return row - 1

    def get_col(self,col):
        """ Converts 1-indexed column to 0-indexed column."""
        return col - 1
    
    def xy_to_excel(self, x, y):
        # x is the column index (1-based)
        # y is the row index (1-based)
        column_letter = get_column_letter(x)
        return f"{column_letter}{y}"
    
    def read_json(self):
        """ Read contents of txt file with 
        JSON format and stores in JSON object."""
        with open(self.json_file, 'r') as json_file:
            string = json_file.read()
            self.json_data = json.loads(string)

    def populate_cells(self):
        """ Populate cells with dummy Cell objects."""
        for row in range(0, len(self.matrix)):
            for col in range(0, len(self.matrix[row])):
                cell_type = "None"
                cell_value = "0"
                self.matrix[row][col] = Cell(cell_type, cell_value, row, col)
    
    def print_matrix(self):
        #print("Matrix: ", self.matrix)
        for row in range(0, len(self.matrix)):
            for col in range(0, len(self.matrix[row])):
                print(self.matrix[row][col], end=", ")
    
    def get_table_data(self, table):
        columns = table.get("Columns", [])
        row_names = table.get("Row Names", [])
        last_row = table.get("Last Row", [])
        return columns, row_names, last_row

    def get_global_col_coordinates(self, columns):
        for row in range(1, self.max_row + 1):
            for col in range(1, self.max_col + 1):
                cell_value = self.ws.cell(row=row, column=col).value
                cell_value = str(cell_value)
                if cell_value in columns:
                    self.global_col_coordinates.append((self.get_row(row), self.get_col(col)))
    
    def pre_pass(self, columns, row_names, last_row):
        """ Store arrays for each table that provides just their row and column information."""
        local_column_occurrences = []
        local_row_name_occurrences = []
        local_last_row_occurrences = []

        #find occurrences of each column in workbook and mark corresponding location in matrix as found column
        for column in columns:
            for row in range(1, self.max_row + 1):
                for col in range(1, self.max_col + 1):
                    cell_value = self.ws.cell(row=row, column=col).value
                    cell_value = str(cell_value)
                    if cell_value == column:
                        local_column_occurrences.append((self.get_row(row), self.get_col(col)))
                        self.matrix[self.get_row(row)][self.get_col(col)].type = "Column"
                        self.matrix[self.get_row(row)][self.get_col(col)].value = column
        
        #find occurrences of each row name in workbook and mark corresponding location in matrix as found row name
        for row_name in row_names:
            for row in range(1, self.max_row + 1):
                for col in range(1, self.max_col + 1):
                    cell_value = self.ws.cell(row=row, column=col).value
                    cell_value = str(cell_value)
                    if cell_value == row_name:
                        local_row_name_occurrences.append((self.get_row(row),self.get_col(col)))
                        self.matrix[self.get_row(row)][self.get_col(col)].type = "Row Name"
                        self.matrix[self.get_row(row)][self.get_col(col)].value = row_name
        
        #find occurrences of last row in workbook and mark corresponding location in matrix as found last row
        for last_row_value in last_row:
            for row in range(1, self.max_row + 1):
                for col in range(1, self.max_col + 1):
                    cell_value = self.ws.cell(row=row, column=col).value
                    cell_value = str(cell_value)
                    if cell_value == '850.00' or cell_value == '$850.00' or cell_value == '850':
                        print(f"*************CELL VALUE*************: {cell_value}")
                    if cell_value == last_row_value:
                        local_last_row_occurrences.append((self.get_row(row),self.get_col(col)))
                        self.matrix[self.get_row(row)][self.get_col(col)].type = "Last Row"
                        self.matrix[self.get_row(row)][self.get_col(col)].value = last_row_value
        # print(f"local_column_occurrences: {local_column_occurrences} \n")
        # print(f"local_row_name_occurrences: {local_row_name_occurrences} \n")
        # print(f"local_last_row_occurrences: {local_last_row_occurrences} \n")
        return local_column_occurrences, local_row_name_occurrences, local_last_row_occurrences
        
    
    def filter_out_floating_rows(self, local_column_occurrences, local_last_row_occurrences):
        """ Filter out floating rows from the table."""
        # Given the last_row_occurrences list, starting from the first coordinate in the last_row occurrence traverse the matrix
        # from that coordinate up (same column) until you reach a row/col that contains a column value (from global column coordinates).
        # If that coordinate also exists in the local column occurrence, then that row value is good and we should keep it
        # If that coordinate does not exist in the local column occurrence, then that row value is bad and we should discard
            # Discard the coordinate of that row value from the last_row_occurrences list
        # Repeat until you reach the end of the local last_row_occurrences list
        filtered_last_row_occurrences = []
        #print(f"local_column_occurrences: {local_column_occurrences} \n")
        for coord in local_last_row_occurrences:
            col_index = coord[1]
            for row_index in range(coord[0], -1, -1):
                if (row_index, col_index) in self.global_col_coordinates and (row_index, col_index) not in local_column_occurrences: 
                    break
                if (row_index, col_index) in local_column_occurrences:
                    filtered_last_row_occurrences.append(coord)
                    break
        print(f"filtered last row occurrences: {filtered_last_row_occurrences} \n")
        return filtered_last_row_occurrences
    
    def filter_out_floating_cols(self, local_column_occurrences, filtered_last_row_occurrences, row_name_occurrences):
        """ Filter out floating columns from the table."""
        # Given the local_column occurrences list, starting from the first coordinate in the the list traverse the matrix
        # from that coordinate down. If a coordinate also exists in the filtered_last_row_occurrences list, then that column value is good and we
        # should keep it.
        # If a coordinate does not exist in the filtered_last_row_occurrences list, then that column value is bad and we should discard
            # Discard the coordinate of that column value from the local_column_occurrences list
        # Repeat until you reach the end of the local_column_occurrences list
        #pdb.set_trace()
        filtered_column_occurrences = []
        for col_occurrence in local_column_occurrences:
            col_index = col_occurrence[1]
            for row_index in range(col_occurrence[0], self.max_row):
                if (row_index, col_index) in filtered_last_row_occurrences or (row_index, col_index) in row_name_occurrences:
                    filtered_column_occurrences.append(col_occurrence)
                    break
        print(f"filtered column occurrences: {filtered_column_occurrences} \n")
        return filtered_column_occurrences
    
    def filter_out_near_identical_tables(self, filtered_column_occurrences, filtered_last_row_occurrences, row_name_occurrences, 
                                        columns, row_names, last_row):
        """ Filter out near identical tables from the table."""
        # Create three lists- one holds all coordinates of columns found so far, one holds all coordinates of row names found so far,
        # and one holds all coordinates of last rows found so far.
        # Given the columns list, starting from the first actual column name, we want to find the first match in our matrix. Once that match is found,
        # we will break our loop, add it to our column coordinates found so far, and add it to our table struct 
        # then find the next first match
        #pdb.set_trace()
        columns_found = []
        row_names_found = []
        last_rows_found = []

        # holds all tables found so far and will return only table with the closest proportion of match to 1
        all_tables_found = False
        tables_found = []
        #Find first match in matrix of columns
        while not all_tables_found:
            # for current table
            current_table_columns = []
            current_table_row_names = []
            current_table_last_row = []
            #Traverse all columns in matrix and find first match and upload it to tablestruct
            first_coordinate_x = -1
            for coordinate in filtered_column_occurrences:
                if coordinate not in self.columns_found:
                    if first_coordinate_x == -1:
                        first_coordinate_x = coordinate[0]
                    row = coordinate[0]
                    col = coordinate[1]
                    cell_value = self.matrix[row][col].value
                    cell_type = self.matrix[row][col].type
                    if cell_value in columns and cell_type == "Column" and row >= first_coordinate_x - 1 and row <= first_coordinate_x + 1\
                        and cell_value not in columns_found:
                        self.columns_found.append(coordinate)
                        columns_found.append(cell_value)
                        current_table_columns.append((row, col))
            first_coordinate_y = -1
            for coordinate in row_name_occurrences:
                if coordinate not in self.row_names_found:
                    if first_coordinate_y == -1:
                        first_coordinate_y = coordinate[1]
                    row = coordinate[0]
                    col = coordinate[1]
                    cell_value = self.matrix[row][col].value
                    cell_type = self.matrix[row][col].type
                    if cell_value in row_names and cell_type == "Row Name" and col >= first_coordinate_y - 1 and col <= first_coordinate_y + 1\
                        and cell_value not in row_names_found:
                        self.row_names_found.append(coordinate)
                        row_names_found.append(cell_value)
                        current_table_row_names.append((row, col))

            for coordinate in filtered_last_row_occurrences:
                #pdb.set_trace()
                row = coordinate[0]
                col = coordinate[1]
                cell_value = self.matrix[row][col].value
                cell_type = self.matrix[row][col].type
                if cell_value in last_row and cell_type == "Last Row":
                    not_floater_row = False
                    for row_index in range(row, -1, -1):
                        if self.matrix[row_index][col].type == "Column" and (row_index, col) in current_table_columns:
                            not_floater_row = True
                            break
                    if not_floater_row:
                        current_table_last_row.append((row, col))
                        last_rows_found.append(cell_value)
            #Given all coordinate lists, process table
            t = Table()
            t.column_headers = current_table_columns.copy()
            t.row_names = current_table_row_names.copy()
            t.last_row = current_table_last_row.copy()
            if len(t.column_headers) + len(t.row_names) + len(t.last_row) > 0:
                tables_found.append(t)
            else:
                all_tables_found = True
            
            # columns_sum = 0
            # row_names_sum = 0
            # last_row_sum = 0
            # for table in tables_found:
            #     columns_sum += len(table.column_headers)
            #     row_names_sum += len(table.row_names)
            #     last_row_sum += len(table.last_row)
            # if len(columns) + len(row_names) + len(last_row) >= columns_sum + row_names_sum + last_row_sum:
            #     all_tables_found = True
            
        total_terms = len(columns) + len(row_names) + len(last_row)
        proportions_list = []
        closest_proportion_tables = []
        for table in tables_found:
            total_terms_in_current_table = len(table.column_headers) + len(table.row_names) + len(table.last_row)
            proportion = total_terms_in_current_table / total_terms
            proportions_list.append(abs(1-proportion))
            # print(f"Pre-Closest Proportion Table:")    
            # print(f"Column Coordinate: {table.column_headers}")
            # print(f"Row Name Coordinate: {table.row_names}")
            # print(f"Row Coordinate: {table.last_row}")
            # print("\n")
        
        if proportions_list:
            min_proportion_value = min(proportions_list)
            for index in range(0, len(proportions_list)):
                if proportions_list[index] == min_proportion_value:
                    closest_proportion_tables.append(tables_found[index])

        for table in closest_proportion_tables:
            last_row_extension = []
            row = table.last_row[0][0]
            col = table.last_row[0][1]
            for col_index in range (col, min(col+len(columns), self.max_col)):
                for row_index in range(row, -1, -1):
                    if self.matrix[row_index][col_index].type == "Column" and self.matrix[row_index][col_index].value in columns:
                        last_row_extension.append((row, col_index))
                        break
            for col_index in range(col, max(-1, col-len(columns)), -1):
                for row_index in range(row, -1, -1):
                    if self.matrix[row_index][col_index].type == "Column" and self.matrix[row_index][col_index].value in columns:
                        last_row_extension.append((row, col_index))
                        break
            table.last_row.extend(last_row_extension)
            index = 0
            print(f"Table: {index}")    
            print(f"Column Coordinates: {table.column_headers}")
            print(f"Row Name Coordinates: {table.row_names}")
            print(f"Row Coordinate: {table.last_row}")
            print("\n")
            index += 1
        return closest_proportion_tables

    def display_table_coordinates(self):
        index = 0
        for table in self.tables:
            all_coords = table.column_headers + table.row_names + table.last_row
            if all_coords:
                # Find the top-left most coordinate (minimum row and column)
                top_left = min(all_coords, key=lambda x: (x[0], x[1]))
                top_left_xl = self.xy_to_excel(top_left[1] + 1, top_left[0] + 1)
                
                # Find the bottom-left most coordinate (maximum row, minimum column)
                bottom_left = min([(coord[0], coord[1]) for coord in all_coords if coord[0] == max(all_coords, key=lambda x: x[0])[0]], key=lambda x: x[1])
                bottom_left_xl = self.xy_to_excel(bottom_left[1] + 1, bottom_left[0] + 1)
                
                # Find the top-right most coordinate (minimum row, maximum column)
                top_right_row = top_left[0]
                top_right_col = max([coord[1] for coord in table.last_row + table.row_names]
)
                top_right_xl = self.xy_to_excel(top_right_col + 1, top_right_row + 1)

                bottom_right_xl = self.xy_to_excel(top_right_col + 1, bottom_left[0] + 1)


                self.corner_coordinates.append((self.ws[top_left_xl].value, self.ws[bottom_left_xl].value, self.ws[top_right_xl].value, self.ws[bottom_right_xl].value))
                print(f"Table {index}:")    
                print(f"Top Left Coordinate: {top_left_xl}, Actual Value: {self.ws[top_left_xl].value}")
                print(f"Bottom Left Coordinate: {bottom_left_xl}, Actual Value: {self.ws[bottom_left_xl].value}")
                print(f"Top Right Coordinate: {top_right_xl}, Actual Value: {self.ws[top_right_xl].value}")
                print(f"Bottom Right Coordinate: {bottom_right_xl}, Actual Value: {self.ws[bottom_right_xl].value}")
                print("\n")
                index += 1
                


    def process_tables(self):
        self.read_json()
        self.populate_cells()
        global_columns = []
        for _, table_data in self.json_data.items():
            columns, row_names, last_row = self.get_table_data(table_data)
            global_columns.extend(columns)
        self.get_global_col_coordinates(global_columns)
        for table_name, table_data in self.json_data.items():
            print(f"Table Name: {table_name}")
            columns, row_names, last_row = self.get_table_data(table_data)
            print(f"With columns of {columns}")
            print(f"With row names of {row_names}")
            print(f"With last rows of {last_row}")
            local_column_occurrences, local_row_name_occurrences, local_last_row_occurrences = self.pre_pass(columns, row_names, last_row)
            # self.print_matrix()
            print(f"Local Column Occurrences: {local_column_occurrences} \n")
            print(f"Local Row Name Occurrences: {local_row_name_occurrences} \n")
            print(f"Local Last Row Occurrences: {local_last_row_occurrences} \n")
            filtered_last_row_occurrences = self.filter_out_floating_rows(local_column_occurrences,local_last_row_occurrences)
            filtered_column_occurrences = self.filter_out_floating_cols(local_column_occurrences, filtered_last_row_occurrences, local_row_name_occurrences)
            filtered_tables = self.filter_out_near_identical_tables(filtered_column_occurrences, filtered_last_row_occurrences,
                                                                   local_row_name_occurrences, columns, row_names, last_row)
            self.tables.extend(filtered_tables)
        self.display_table_coordinates()
        

def main():
    ex = ExcelProcessor(
        xlsx_file=argv[1],
        json_file=argv[2]
    )
    ex.print_workbook_contents()
    ex.process_tables()

if __name__ == "__main__":
    main()

