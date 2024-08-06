import openpyxl
from openpyxl.utils import get_column_letter
from collections import Counter, deque
import json
from sys import argv
from math import inf
import pandas as pd
import pdb

# Helper class to represent a table in the spreadsheet
class Table:
    def __init__(self):
        self.column_headers = []
        self.row_name_vals = []
        self.last_row_vals = []
        self.row_names_found = []
        self.last_row_found = []
        self.last_row_name_val = []
        self.last_row_val = []
        self.corner_coords = []
        self.num_columns = 0


# Helper class to represent a cell in the spreadsheet
class Cell:
    def __init__(self, type, value, row, col):
        self.type = type # "None", "Column", "Row Name", "Last Row"
        self.value = value # self.ws.cell(row=self.row, column=self.col).value
        self.row = row
        self.col = col
    def __repr__(self):
        return f"{self.type}: {self.value}"

# Main Excel Processing Class
class ExcelProcessor:
    def __init__(self, xlsx_file, json_file):
        self.wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        self.ws = self.wb.active
        self.json_file = json_file
        self.max_row = self.ws.max_row
        self.max_col = self.ws.max_column
        self.matrix =  [[None for _ in range(self.max_col)] for _ in range(self.max_row)]
        self.json_data = None
        self.global_coordinates_found = []
        self.global_columns_found = []
        self.tables = []
        # for all such coordinates found
        self.columns_found = []
        self.row_names_found = []
        self.last_row_found = []
        self.corner_coordinates = []


    def print_workbook_contents(self):
        # Iterate through each row in the worksheet
        for row in self.ws.iter_rows(values_only=True):
            # Print each cell value in the row
            print(row)
        
    def get_row(self,row):
        """ Converts 1-indexed row to 0-indexed row."""
        return row - 1

    def get_col(self,col):
        """ Converts 1-indexed column to 0-indexed column."""
        return col - 1
    
    def xy_to_excel(self, x, y):
        # x is the column index (1-based)
        # y is the row index (1-based)
        column_letter = get_column_letter(x)
        return f"{column_letter}{y}"
    
    def read_json(self):
        """ Read contents of txt file with 
        JSON format and stores in JSON object."""
        with open(self.json_file, 'r') as json_file:
            string = json_file.read()
            self.json_data = json.loads(string)

    def populate_cells(self):
        """ Populate cells with dummy Cell objects."""
        for row in range(0, len(self.matrix)):
            for col in range(0, len(self.matrix[row])):
                cell_type = "None"
                cell_value = "0"
                self.matrix[row][col] = Cell(cell_type, cell_value, row, col)
    
    def print_matrix(self):
        #print("Matrix: ", self.matrix)
        for row in range(0, len(self.matrix)):
            for col in range(0, len(self.matrix[row])):
                print(self.matrix[row][col], end=", ")
    
    def get_table_data(self, table):
        columns = table.get("Columns", [])
        row_names = table.get("Row Names", [])
        last_row = table.get("Last Row", [])
        return columns, row_names, last_row
    
    def pre_pass(self, columns, row_names, last_row):
        """ Store arrays for each table that provides just their row and column information."""
        local_column_occurrences = []
        local_row_name_occurrences = []
        local_last_row_occurrences = []

        #find occurrences of each column in workbook and mark corresponding location in matrix as found column
        for column in columns:
            for row in range(1, self.max_row + 1):
                for col in range(1, self.max_col + 1):
                    cell_value = self.ws.cell(row=row, column=col).value
                    cell_value = str(cell_value)
                    if cell_value == column:
                        local_column_occurrences.append((self.get_row(row), self.get_col(col)))
                        self.matrix[self.get_row(row)][self.get_col(col)].type = "Column"
                        self.matrix[self.get_row(row)][self.get_col(col)].value = column
        
        #find occurrences of each row name in workbook and mark corresponding location in matrix as found row name
        for row_name in row_names:
            for row in range(1, self.max_row + 1):
                for col in range(1, self.max_col + 1):
                    cell_value = self.ws.cell(row=row, column=col).value
                    cell_value = str(cell_value)
                    if cell_value == row_name:
                        local_row_name_occurrences.append((self.get_row(row),self.get_col(col)))
                        self.matrix[self.get_row(row)][self.get_col(col)].type = "Row Name"
                        self.matrix[self.get_row(row)][self.get_col(col)].value = row_name
        
        #find occurrences of last row in workbook and mark corresponding location in matrix as found last row
        for last_row_value in last_row:
            for row in range(1, self.max_row + 1):
                for col in range(1, self.max_col + 1):
                    cell_value = self.ws.cell(row=row, column=col).value
                    cell_value = str(cell_value)
                    if cell_value == last_row_value:
                        local_last_row_occurrences.append((self.get_row(row),self.get_col(col)))
                        self.matrix[self.get_row(row)][self.get_col(col)].type = "Last Row"
                        self.matrix[self.get_row(row)][self.get_col(col)].value = last_row_value
        # print(f"local_column_occurrences: {local_column_occurrences} \n")
        # print(f"local_row_name_occurrences: {local_row_name_occurrences} \n")
        # print(f"local_last_row_occurrences: {local_last_row_occurrences} \n")
        return local_column_occurrences, local_row_name_occurrences, local_last_row_occurrences

    def find_column_sequence(self, local_column_occurrences, column_names):
        """ From the local column occurrences, find each sequence of columnns available"""
        column_sequences = []
        columns_seen = []

        #Pass 0: For each row, note where a column sequence occurs
        for coordinate in local_column_occurrences:
            if coordinate in columns_seen or coordinate in self.global_columns_found:
                continue
            else:
                columns_counter = Counter(column_names)
                row = coordinate[0]
                starting_col = coordinate[1]
                current_sequence = []
                for col in range(starting_col, self.max_col):
                    if self.matrix[row][col].type == "Column" and self.matrix[row][col].value in column_names:
                        if columns_counter[self.matrix[row][col].value] > 0:
                            columns_counter[self.matrix[row][col].value] -= 1
                            current_sequence.append((row, col))
                columns_seen.extend(current_sequence)
                column_sequences.append(current_sequence)
                current_sequence = []
        
        column_sequences.sort(key=lambda seq: seq[0][0] if seq else 0)
        print(f"Column Sequences: {column_sequences} \n")
        if not column_sequences:
            return None

        # Pass 1: Find the sequence of column headers with the highest match
        best_sequence_score = inf
        best_sequence = []
        global_sequence_set = set(self.global_columns_found)
        for sequence in column_sequences:
            #check if sequence is already a column found above
            # Check if any item in the subsequence is in the sequence set
            if any(item in global_sequence_set for item in sequence):
                continue
            sequence_score = len(sequence) / len(column_names)
            inverted_sequence_score = abs((1 - sequence_score))
            if inverted_sequence_score < best_sequence_score:
                best_sequence_score = inverted_sequence_score
                best_sequence = sequence
        
        if best_sequence_score <= abs(1 - (len(column_names) - 1) / (len(column_names))):
            self.global_columns_found.extend(best_sequence)
            return best_sequence
        
        #Pass 2: We've computed sequence scores for each sequence and noted in which row each sequence lies
                 # If passing threshold sequence DNE, we need to examine adjacent rows max 2 at a time to find if they contain the entire sequence
        best_sequence_score = inf
        best_sequence = []
        for index, sequence in enumerate(column_sequences):
            if "PROJECTED MONTHLY INCOME" in column_names:
                #pdb.set_trace()
                pass
            if not sequence:
                continue
            if any(item in global_sequence_set for item in sequence):
                continue
            for next_sequence in column_sequences[index + 1:]:
                if not next_sequence:
                    continue
                if next_sequence[0][0] == sequence[0][0] + 1:
                    column_counter = Counter(column_names)
                    combined_sequence = sequence + next_sequence
                    for item in combined_sequence:
                        if column_counter[self.matrix[item[0]][item[1]].value] > 0:
                            column_counter[self.matrix[item[0]][item[1]].value] -= 1
                    found_sequence = True
                    for item, count in column_counter.items():
                        if count > 0:
                            found_sequence = False
                            break
                    if found_sequence:
                        self.global_columns_found.extend(combined_sequence)
                        return combined_sequence
        
        #Pass 3: If we still haven't found a sequence that satisfies the threshold, we need to assume a subtable exists and construct a sequence given a main column header
        # Reconstruct the column sequence from self.global_columns_found
        old_column_sequences = column_sequences.copy()
        old_column_sequences.sort(key=lambda seq: seq[0][0] if seq else 0)
        print(old_column_sequences)
        min_value = 0
        #Minor fix for when the sequeence is not found altogether
        if not old_column_sequences[0]:
            pass
        elif old_column_sequences[0][0]:
            min_value = old_column_sequences[0][0][0]
        column_sequences = []
        for item in self.global_columns_found:
            columns_counter = Counter(column_names)
            row = item[0]
            starting_col = item[1]
            current_sequence = []
            for col in range(starting_col, self.max_col):
                if self.matrix[row][col].type == "Column" and self.matrix[row][col].value in column_names:
                    if columns_counter[self.matrix[row][col].value] > 0:
                        columns_counter[self.matrix[row][col].value] -= 1
                        current_sequence.append((row, col))
            columns_seen.extend(current_sequence)
            column_sequences.append(current_sequence)
            current_sequence = []
        
        highest_sequence = []
        if column_sequences:
            column_sequences.sort(key=lambda seq: seq[0][0] if seq else 0)
            largest_value = 0
            for sequence in column_sequences:
                if not sequence:
                    continue
                if sequence[0][0] > largest_value and sequence[0][0] < min_value:
                    largest_value = sequence[0][0]
                    highest_sequence = sequence
        
        final_sequence = []
        old_column_sequences.append(highest_sequence)
        for sequence in old_column_sequences:
            if not sequence:
                continue
            for item in sequence:
                final_sequence.append(item)
        return final_sequence
    
    def form_table_with_column_sequence(self, column_sequence):
        table = Table()
        table.column_headers = column_sequence
        return table
    
    def find_column_boundaries(self, column_sequence):
        col_sequence = column_sequence.copy()
        if not col_sequence:
            return None
        highest_row_value = max(col_sequence, key=lambda x: x[0])[0]
        lowest_col_value = min(col_sequence, key=lambda x: x[1])[1]
        highest_col_value = max(col_sequence, key=lambda x: x[1])[1]
        column_boundary_left = (highest_row_value, lowest_col_value)
        column_boundary_right = (highest_row_value, highest_col_value)
        return column_boundary_left, column_boundary_right
        
    
    def find_final_row_name(self, local_row_name_occurrences, row_names, column_boundary_left):
        if local_row_name_occurrences:
            next_table_found = False
            if "Electricity" in row_names:
                #pdb.set_trace()
                pass
            row_name_sequence = []
            row_name_counter = Counter(row_names)
            row_index = column_boundary_left[0] + 1

            while not next_table_found:
                column_boundary_left_col = column_boundary_left[1] - 1
                if column_boundary_left[1] - 1 < 0:
                    column_boundary_left_col = column_boundary_left[1]
                for col in range(column_boundary_left_col, column_boundary_left_col + 2):
                    if "Gym Membership" in row_names:
                        #pdb.set_trace()
                        pass
                    if any((row_index, col) == coord for coord in self.global_columns_found) or row_index == self.max_row:
                        next_table_found = True
                    elif self.matrix[row_index][col].type == "Row Name" and self.matrix[row_index][col].value in row_names:
                        row_name_counter[self.matrix[row_index][col].value] -= 1
                        row_name_sequence.append((row_index, col))
                row_index += 1
            if row_name_sequence:
                return max(row_name_sequence, key=lambda x: x[0])
        else:
            return None
    
    def find_final_last_row(self, local_last_row_occurrences, last_row, row_names, column_boundary_left, column_boundary_right):
        if local_last_row_occurrences:
            next_table_found = False
            row_index = column_boundary_left[0] + 1
            while not next_table_found:
                for col in range(column_boundary_left[1], column_boundary_right[1]):
                    if "Life Insurance" in last_row:
                        #pdb.set_trace()
                        pass
                    if row_index == self.max_row:
                        next_table_found = True
                    elif any((row_index, col) == coord for coord in self.global_columns_found) or \
                    (self.matrix[row_index][col].type == "Last Row" and self.matrix[row_index][col].value not in last_row) or\
                    (self.matrix[row_index][col].type == "Row Name" and self.matrix[row_index][col].value not in row_names):
                        next_table_found = True
                row_index += 1
            row_index -= 1
            if row_index == self.max_row:
                row_index -= 1
            for row in range(row_index, column_boundary_left[0], -1):
                for col in range(column_boundary_left[1], column_boundary_right[1]):
                    if self.matrix[row][col].type == "Last Row" \
                    or self.matrix[row][col].type == "Row Name" \
                        and self.matrix[row][col].value in last_row or self.matrix[row][col].value in row_names:
                        return (row, col)
        else:
            return None
    
    def find_table_corners(self, table):
        column_boundary_left, column_boundary_right = self.find_column_boundaries(table.column_headers)
        top_left = column_boundary_left
        top_left_to_excel = self.xy_to_excel(top_left[1] + 1, top_left[0] + 1)
        top_right = column_boundary_right
        top_right_to_excel = self.xy_to_excel(top_right[1] + 1, top_right[0] + 1)
        if table.last_row_val:
            bottom_left = (table.last_row_val[0], column_boundary_left[1])
            bottom_left_to_excel = self.xy_to_excel(bottom_left[1] + 1, bottom_left[0] + 1)
            bottom_right = (table.last_row_val[0], column_boundary_right[1])
            bottom_right_to_excel = self.xy_to_excel(bottom_right[1] + 1, bottom_right[0] + 1)
        elif table.last_row_name_val:
            bottom_left = (table.last_row_name_val[0], column_boundary_left[1])
            bottom_left_to_excel = self.xy_to_excel(bottom_left[1] + 1, bottom_left[0] + 1)
            bottom_right = (table.last_row_name_val[0], column_boundary_right[1])
            bottom_right_to_excel = self.xy_to_excel(bottom_right[1] + 1, bottom_right[0] + 1)
        else:
            return None
        print(f"Top Left: {top_left_to_excel}, Value: {self.ws[top_left_to_excel].value}")
        print(f"Top Right: {top_right_to_excel}, Value: {self.ws[top_right_to_excel].value}")
        print(f"Bottom Left: {bottom_left_to_excel}, Value: {self.ws[bottom_left_to_excel].value}")
        print(f"Bottom Right: {bottom_right_to_excel}, Value: {self.ws[bottom_right_to_excel].value}")
        return top_left_to_excel, top_right_to_excel, bottom_left_to_excel, bottom_right_to_excel
    
    #def display_table_as_df(self,table):
    def find_df_table(self, table):
        """ Create pandas DataFrames for each table using corner coordinates. """
        if table.corner_coords:
            top_left, top_right, bottom_left, bottom_right = table.corner_coords
            start_row = int(top_left[1:])
            end_row = int(bottom_left[1:])
            start_col = openpyxl.utils.column_index_from_string(top_left[0])
            end_col = openpyxl.utils.column_index_from_string(top_right[0])
            
            data = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    cell_value = self.ws.cell(row=row, column=col).value
                    row_data.append(cell_value)
                data.append(row_data)
            
            df = pd.DataFrame(data)
            print(f"DF: \n")
            print(f"{df}")
            print("\n")

            
    def process_tables(self):
        self.read_json()
        self.populate_cells()
        # global_columns = []
        # for _, table_data in self.json_data.items():
        #     columns, row_names, last_row = self.get_table_data(table_data)
        #     global_columns.extend(columns)
        for table_name, table_data in self.json_data.items():
            print(f"Table Name: {table_name}")
            columns, row_names, last_row = self.get_table_data(table_data)
            num_columns = len(columns)
            # print(f"With columns of {columns}")
            # print(f"With row names of {row_names}")
            # print(f"With last rows of {last_row}")
            local_column_occurrences, local_row_name_occurrences, local_last_row_occurrences = self.pre_pass(columns, row_names, last_row)
            # self.print_matrix()
            # print(f"Local Column Occurrences: {local_column_occurrences} \n")
            # print(f"Local Row Name Occurrences: {local_row_name_occurrences} \n")
            # print(f"Local Last Row Occurrences: {local_last_row_occurrences} \n")
            column_sequence = self.find_column_sequence(local_column_occurrences, columns)
            # print(f"Columns Found: {column_sequence} \n")
            filtered_table = self.form_table_with_column_sequence(column_sequence)
            filtered_table.row_names_found = local_row_name_occurrences
            filtered_table.last_row_found = local_last_row_occurrences
            filtered_table.row_name_vals = row_names
            filtered_table.last_row_vals = last_row
            self.tables.append(filtered_table)
        for index, table in enumerate(self.tables):
            if table.column_headers:
                column_boundary_left, column_boundary_right = self.find_column_boundaries(table.column_headers)
                if table.row_names_found:
                    table.last_row_name_val = self.find_final_row_name(table.row_names_found, table.row_name_vals, column_boundary_left)
                if table.last_row_found:
                    table.last_row_val = self.find_final_last_row(table.last_row_found, table.last_row_vals, table.row_name_vals, column_boundary_left, column_boundary_right)
                table.corner_coords = self.find_table_corners(table)
                # print(f"Table {index}:")
                # print(f"Column Headers: {table.column_headers}")
                # print(f"Last Row Name: {table.last_row_name_val}")
                # print(f"Last Row: {table.last_row_val}")
                self.find_df_table(table)



        #self.display_table_coordinates()

def main():
    ex = ExcelProcessor(
        xlsx_file=argv[1],
        json_file=argv[2]
    )
    ex.process_tables()

if __name__ == "__main__":
    main()                      


