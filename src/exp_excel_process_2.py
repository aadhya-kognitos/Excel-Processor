import openpyxl
from openpyxl.utils import get_column_letter
import json
from sys import argv
from math import inf
import pdb

# Helper class to represent a table in the spreadsheet
class Table:
    def __init__(self):
        self.column_headers = []
        self.row_names = []
        self.last_row = []
        self.corner_coords = []
        self.num_columns = 0


# Helper class to represent a cell in the spreadsheet
class Cell:
    def __init__(self, type, value, row, col):
        self.type = type # "None", "Column", "Row Name", "Last Row"
        self.value = value # self.ws.cell(row=self.row, column=self.col).value
        self.row = row
        self.col = col
    def __repr__(self):
        return f"{self.type}: {self.value}"

# Main Excel Processing Class
class ExcelProcessor:
    def __init__(self, xlsx_file, json_file):
        self.wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        self.ws = self.wb.active
        self.json_file = json_file
        self.max_row = self.ws.max_row
        self.max_col = self.ws.max_column
        self.matrix =  [[None for _ in range(self.max_col)] for _ in range(self.max_row)]
        self.json_data = None
        self.global_coordinates_found = []
        self.tables = []
        # for all such coordinates found
        self.columns_found = []
        self.row_names_found = []
        self.last_row_found = []
        self.corner_coordinates = []


    def print_workbook_contents(self):
        # Iterate through each row in the worksheet
        for row in self.ws.iter_rows(values_only=True):
            # Print each cell value in the row
            print(row)
        
    def get_row(self,row):
        """ Converts 1-indexed row to 0-indexed row."""
        return row - 1

    def get_col(self,col):
        """ Converts 1-indexed column to 0-indexed column."""
        return col - 1
    
    def xy_to_excel(self, x, y):
        # x is the column index (1-based)
        # y is the row index (1-based)
        column_letter = get_column_letter(x)
        return f"{column_letter}{y}"
    
    def read_json(self):
        """ Read contents of txt file with 
        JSON format and stores in JSON object."""
        with open(self.json_file, 'r') as json_file:
            string = json_file.read()
            self.json_data = json.loads(string)

    def populate_cells(self):
        """ Populate cells with dummy Cell objects."""
        for row in range(0, len(self.matrix)):
            for col in range(0, len(self.matrix[row])):
                cell_type = "None"
                cell_value = "0"
                self.matrix[row][col] = Cell(cell_type, cell_value, row, col)
    
    def print_matrix(self):
        #print("Matrix: ", self.matrix)
        for row in range(0, len(self.matrix)):
            for col in range(0, len(self.matrix[row])):
                print(self.matrix[row][col], end=", ")
    
    def get_table_data(self, table):
        columns = table.get("Columns", [])
        row_names = table.get("Row Names", [])
        last_row = table.get("Last Row", [])
        return columns, row_names, last_row
    
    def pre_pass(self, columns, row_names, last_row):
        """ Store arrays for each table that provides just their row and column information."""
        local_column_occurrences = []
        local_row_name_occurrences = []
        local_last_row_occurrences = []

        #find occurrences of each column in workbook and mark corresponding location in matrix as found column
        for column in columns:
            for row in range(1, self.max_row + 1):
                for col in range(1, self.max_col + 1):
                    cell_value = self.ws.cell(row=row, column=col).value
                    cell_value = str(cell_value)
                    if cell_value == column:
                        local_column_occurrences.append((self.get_row(row), self.get_col(col)))
                        self.matrix[self.get_row(row)][self.get_col(col)].type = "Column"
                        self.matrix[self.get_row(row)][self.get_col(col)].value = column
        
        #find occurrences of each row name in workbook and mark corresponding location in matrix as found row name
        for row_name in row_names:
            for row in range(1, self.max_row + 1):
                for col in range(1, self.max_col + 1):
                    cell_value = self.ws.cell(row=row, column=col).value
                    cell_value = str(cell_value)
                    if cell_value == row_name:
                        local_row_name_occurrences.append((self.get_row(row),self.get_col(col)))
                        self.matrix[self.get_row(row)][self.get_col(col)].type = "Row Name"
                        self.matrix[self.get_row(row)][self.get_col(col)].value = row_name
        
        #find occurrences of last row in workbook and mark corresponding location in matrix as found last row
        for last_row_value in last_row:
            for row in range(1, self.max_row + 1):
                for col in range(1, self.max_col + 1):
                    cell_value = self.ws.cell(row=row, column=col).value
                    cell_value = str(cell_value)
                    if cell_value == last_row_value:
                        local_last_row_occurrences.append((self.get_row(row),self.get_col(col)))
                        self.matrix[self.get_row(row)][self.get_col(col)].type = "Last Row"
                        self.matrix[self.get_row(row)][self.get_col(col)].value = last_row_value
        # print(f"local_column_occurrences: {local_column_occurrences} \n")
        # print(f"local_row_name_occurrences: {local_row_name_occurrences} \n")
        # print(f"local_last_row_occurrences: {local_last_row_occurrences} \n")
        return local_column_occurrences, local_row_name_occurrences, local_last_row_occurrences
        
    
    def find_last_row(self, local_last_row_occurrences, local_last_row_names):
        """ Find the last row for the given table"""

        #If a coordinate has already been seen (global last row coords), skip it
        #Else
            #Match value in row col to actual column headers list
            #Build out a list containing all of these column headers, until you reach the last value/last cell
            #Update the row with the maximum matches (prop. closest to 1 of matched keywords)
        #filtered last row occurrences is the list of coordinates with most normalized match
        potential_lst = []
        local_coords_found = []
        potential_lst_score = inf
        for coord in local_last_row_occurrences:
            if coord == (22, 1) or coord == (21, 6) or coord == (22,4):
                #pdb.set_trace()
                print(coord)

            if (coord[0], coord[1]) in self.global_coordinates_found:
                continue
            else:
                #Create a list to store current last row values seen
                last_row_lst = []
                #To keep track of how many last row values we have actually seen in current sequence
                last_rows_seen = 0
                #Loop through every row name in the current table
                for row_name in local_last_row_names:
                    #Store the x-coordinate as we are operating on that row
                    row_val = coord[0]
                    #For every coordinate in the current row
                    for coordinate in range(0, self.max_col):
                        #if the value of the coordinate equals the row name
                        if self.matrix[row_val][coordinate].value == row_name and (row_val, coordinate) not in self.global_coordinates_found:
                            #Append to our current last row values list
                            last_row_lst.append((row_val, coordinate))
                            last_rows_seen += 1
                            #Break as we don't want to include duplicates
                            break
                    #If the total number of last rows seen now equals the length of local_last_row_names, break out
                    if last_rows_seen >= len(local_last_row_names):
                        break
            #pdb.set_trace()
            #Get the score by diving the last rows seen by the length of local_last_row_names
            last_row_lst_score = last_rows_seen / len(local_last_row_names)
            #If the score is less than the potential score, update the potential score and the potential lst
            if abs(1 - last_row_lst_score) < potential_lst_score:
                potential_lst = last_row_lst
                potential_lst_score = abs(1- last_row_lst_score)
        
        filtered_last_row_occurrences = potential_lst.copy()
        self.global_coordinates_found.extend(filtered_last_row_occurrences)
        return filtered_last_row_occurrences
    
    
    def find_row_names(self, local_row_name_occurrences, local_row_name_names):
        """ Finds the row names for a given table."""
        #If a coordinate has already been seen (global last row coords), skip it
        #Else
            #Match value in row col to actual column headers list
            #Build out a list containing all of these column headers, until you reach the last value/last cell
            #Update the row with the maximum matches (prop. closest to 1 of matched keywords)
        #filtered row name occurrences is the list of coordinates with most normalized match
        potential_lst = []
        potential_lst_score = inf
        for coord in local_row_name_occurrences:
            #pdb.set_trace()
            if (coord[0], coord[1]) in self.global_coordinates_found:
                continue
            else:
                #Create a list to store current row name values seen
                row_name_lst = []
                #To keep track of how many row name values we have actually seen in current sequence
                row_names_seen = 0
                #Loop through every row name in the current table
                for row_name in local_row_name_names:
                    #Store the y-coordinate as we are operating on that column
                    col_val = coord[1]
                    #For every coordinate in the current row
                    for coordinate in range(0, self.max_row):
                        #if the value of the coordinate equals the row name
                        if self.matrix[coordinate][col_val].value == row_name and (coordinate, col_val) not in self.global_coordinates_found:
                            #Append to our current row name values list
                            row_name_lst.append((coordinate, col_val))
                            row_names_seen += 1
                            #Break as we don't want to include duplicates
                            break
                    #If the total number of row names seen now equals the length of local_names, break out
                    if row_names_seen >= len(local_row_name_names):
                        break
            #pdb.set_trace()
            #Get the score by diving the last rows seen by the length of local_last_row_names
            row_name_lst_score = row_names_seen / len(local_row_name_names)
            #If the score is less than the potential score, update the potential score and the potential lst
            if abs(1 - row_name_lst_score) < potential_lst_score:
                potential_lst = row_name_lst
                potential_lst_score = abs(1 - row_name_lst_score)
        
        filtered_row_name_occurrences = potential_lst.copy()
        self.global_coordinates_found.extend(filtered_row_name_occurrences)
        return filtered_row_name_occurrences


    def find_column(self, local_column_occurrences, local_column_names):
        """ Finds the columns for a given table."""
        #If a coordinate has already been seen (global last row coords), skip it
        #Else
            #Match value in row col to actual column headers list
            #Build out a list containing all of these column headers, until you reach the last value/last cell
            #Update the row with the maximum matches (prop. closest to 1 of matched keywords)
        #filtered row name occurrences is the list of coordinates with most normalized match
        potential_lst = []
        potential_lst_score = inf
        for coord in local_column_occurrences:
            #pdb.set_trace()
            if (coord[0], coord[1]) in self.global_coordinates_found:
                continue
            else:
                #Create a list to store current column values seen
                column_lst = []
                #To keep track of how many column values we have actually seen in current sequence
                columns_seen = 0
                #Loop through every column name in the current table
                for column_name in local_column_names:
                    #Store the x-coordinate as we are operating on that row
                    row_val = coord[0]
                    #For every coordinate in the current row
                    for coordinate in range(0, self.max_col):
                        #if the value of the coordinate equals the column name
                        if self.matrix[row_val][coordinate].value == column_name and (row_val, coordinate) not in self.global_coordinates_found:
                            #Append to our current column values list
                            column_lst.append((row_val, coordinate))
                            columns_seen += 1
                            #Break as we don't want to include duplicates
                            break
                    #If the total number of column names seen now equals the length of local_column_names, break out
                    if columns_seen >= len(local_column_names):
                        break
            #pdb.set_trace()
            #Get the score by diving the last rows seen by the length of local_last_row_names
            column_lst_score = columns_seen / len(local_column_names)
            #If the score is less than the potential score, update the potential score and the potential lst
            if abs(1 - column_lst_score) < potential_lst_score:
                potential_lst = column_lst
                potential_lst_score = abs(1 - column_lst_score)
        
        filtered_column_occurrences = potential_lst.copy()
        self.global_coordinates_found.extend(filtered_column_occurrences)
        return filtered_column_occurrences
    
    def form_table(self, filtered_column_occurrences, filtered_last_row_occurrences, filtered_row_name_occurrences, num_columns):
        """ Filter out near identical tables from the table."""
        table = Table()
        table.column_headers = filtered_column_occurrences
        table.row_names = filtered_row_name_occurrences
        table.last_row = filtered_last_row_occurrences
        table.num_columns = num_columns
        return table
    

    def display_table_coordinates(self):
        index = 0
        for table in self.tables:
            all_coords = table.column_headers + table.row_names + table.last_row
            if all_coords:
                
                # Find the top-left most coordinate (minimum row and column)
                top_left = min(table.column_headers, key=lambda x: (x[0], x[1]))
                top_left_xl = self.xy_to_excel(top_left[1] + 1, top_left[0] + 1)
                
                # Find the bottom-left most coordinate (maximum row, minimum column)
                bottom_left = min([(coord[0], coord[1]) for coord in all_coords if coord[0] == max(all_coords, key=lambda x: x[0])[0]], key=lambda x: x[1])
                bottom_left = (bottom_left[0], top_left[1])
                bottom_left_xl = self.xy_to_excel(bottom_left[1] + 1, bottom_left[0] + 1)
                
                # Find the top-right most coordinate (minimum row, maximum column)
                top_right_row = top_left[0]
                top_right_col = top_left[1] + table.num_columns - 1
                top_right_xl = self.xy_to_excel(top_right_col + 1, top_right_row + 1)

                bottom_right_xl = self.xy_to_excel(top_right_col + 1, bottom_left[0] + 1)
                
                self.corner_coordinates.append((self.ws[top_left_xl].value, self.ws[bottom_left_xl].value, self.ws[top_right_xl].value, self.ws[bottom_right_xl].value))
                print(f"Table {index}:")    
                print(f"Top Left Coordinate: {top_left_xl}, Actual Value: {self.ws[top_left_xl].value}")
                print(f"Bottom Left Coordinate: {bottom_left_xl}, Actual Value: {self.ws[bottom_left_xl].value}")
                print(f"Top Right Coordinate: {top_right_xl}, Actual Value: {self.ws[top_right_xl].value}")
                print(f"Bottom Right Coordinate: {bottom_right_xl}, Actual Value: {self.ws[bottom_right_xl].value}")
                print("\n")
                index += 1

    def process_tables(self):
        self.read_json()
        self.populate_cells()
        # global_columns = []
        # for _, table_data in self.json_data.items():
        #     columns, row_names, last_row = self.get_table_data(table_data)
        #     global_columns.extend(columns)
        for table_name, table_data in self.json_data.items():
            print(f"Table Name: {table_name}")
            columns, row_names, last_row = self.get_table_data(table_data)
            num_columns = len(columns)
            print(f"With columns of {columns}")
            print(f"With row names of {row_names}")
            print(f"With last rows of {last_row}")
            local_column_occurrences, local_row_name_occurrences, local_last_row_occurrences = self.pre_pass(columns, row_names, last_row)
            # self.print_matrix()
            print(f"Local Column Occurrences: {local_column_occurrences} \n")
            print(f"Local Row Name Occurrences: {local_row_name_occurrences} \n")
            print(f"Local Last Row Occurrences: {local_last_row_occurrences} \n")
            filtered_column_occurrences = self.find_column(local_column_occurrences, columns)
            print(f"Columns Found: {filtered_column_occurrences} \n")
            filtered_row_name_occurrences = self.find_row_names(local_row_name_occurrences, row_names)
            print(f"Row Names Found: {filtered_row_name_occurrences} \n")
            filtered_last_row_occurrences = self.find_last_row(local_last_row_occurrences, last_row)
            print(f"Last Row Found: {filtered_last_row_occurrences} \n")
            filtered_table = self.form_table(filtered_column_occurrences, filtered_last_row_occurrences, filtered_row_name_occurrences, num_columns)
            self.tables.append(filtered_table)
        self.display_table_coordinates()
        

def main():
    ex = ExcelProcessor(
        xlsx_file=argv[1],
        json_file=argv[2]
    )
    ex.process_tables()

if __name__ == "__main__":
    main()

