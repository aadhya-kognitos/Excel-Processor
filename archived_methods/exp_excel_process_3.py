import openpyxl
from openpyxl.utils import get_column_letter
from collections import Counter, deque
import json
from sys import argv
from math import inf
import pdb

# Helper class to represent a table in the spreadsheet
class Table:
    def __init__(self):
        self.column_headers = []
        self.row_names = []
        self.last_row = []
        self.corner_coords = []
        self.num_columns = 0


# Helper class to represent a cell in the spreadsheet
class Cell:
    def __init__(self, type, value, row, col):
        self.type = type # "None", "Column", "Row Name", "Last Row"
        self.value = value # self.ws.cell(row=self.row, column=self.col).value
        self.row = row
        self.col = col
    def __repr__(self):
        return f"{self.type}: {self.value}"

# Main Excel Processing Class
class ExcelProcessor:
    def __init__(self, xlsx_file, json_file):
        self.wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        self.ws = self.wb.active
        self.json_file = json_file
        self.max_row = self.ws.max_row
        self.max_col = self.ws.max_column
        self.matrix =  [[None for _ in range(self.max_col)] for _ in range(self.max_row)]
        self.json_data = None
        self.global_coordinates_found = []
        self.global_columns_found = []
        self.tables = []
        # for all such coordinates found
        self.columns_found = []
        self.row_names_found = []
        self.last_row_found = []
        self.corner_coordinates = []


    def print_workbook_contents(self):
        # Iterate through each row in the worksheet
        for row in self.ws.iter_rows(values_only=True):
            # Print each cell value in the row
            print(row)
        
    def get_row(self,row):
        """ Converts 1-indexed row to 0-indexed row."""
        return row - 1

    def get_col(self,col):
        """ Converts 1-indexed column to 0-indexed column."""
        return col - 1
    
    def xy_to_excel(self, x, y):
        # x is the column index (1-based)
        # y is the row index (1-based)
        column_letter = get_column_letter(x)
        return f"{column_letter}{y}"
    
    def read_json(self):
        """ Read contents of txt file with 
        JSON format and stores in JSON object."""
        with open(self.json_file, 'r') as json_file:
            string = json_file.read()
            self.json_data = json.loads(string)

    def populate_cells(self):
        """ Populate cells with dummy Cell objects."""
        for row in range(0, len(self.matrix)):
            for col in range(0, len(self.matrix[row])):
                cell_type = "None"
                cell_value = "0"
                self.matrix[row][col] = Cell(cell_type, cell_value, row, col)
    
    def print_matrix(self):
        #print("Matrix: ", self.matrix)
        for row in range(0, len(self.matrix)):
            for col in range(0, len(self.matrix[row])):
                print(self.matrix[row][col], end=", ")
    
    def get_table_data(self, table):
        columns = table.get("Columns", [])
        row_names = table.get("Row Names", [])
        last_row = table.get("Last Row", [])
        return columns, row_names, last_row
    
    def pre_pass(self, columns, row_names, last_row):
        """ Store arrays for each table that provides just their row and column information."""
        local_column_occurrences = []
        local_row_name_occurrences = []
        local_last_row_occurrences = []

        #find occurrences of each column in workbook and mark corresponding location in matrix as found column
        for column in columns:
            for row in range(1, self.max_row + 1):
                for col in range(1, self.max_col + 1):
                    cell_value = self.ws.cell(row=row, column=col).value
                    cell_value = str(cell_value)
                    if cell_value == column:
                        local_column_occurrences.append((self.get_row(row), self.get_col(col)))
                        self.matrix[self.get_row(row)][self.get_col(col)].type = "Column"
                        self.matrix[self.get_row(row)][self.get_col(col)].value = column
        
        #find occurrences of each row name in workbook and mark corresponding location in matrix as found row name
        for row_name in row_names:
            for row in range(1, self.max_row + 1):
                for col in range(1, self.max_col + 1):
                    cell_value = self.ws.cell(row=row, column=col).value
                    cell_value = str(cell_value)
                    if cell_value == row_name:
                        local_row_name_occurrences.append((self.get_row(row),self.get_col(col)))
                        self.matrix[self.get_row(row)][self.get_col(col)].type = "Row Name"
                        self.matrix[self.get_row(row)][self.get_col(col)].value = row_name
        
        #find occurrences of last row in workbook and mark corresponding location in matrix as found last row
        for last_row_value in last_row:
            for row in range(1, self.max_row + 1):
                for col in range(1, self.max_col + 1):
                    cell_value = self.ws.cell(row=row, column=col).value
                    cell_value = str(cell_value)
                    if cell_value == last_row_value:
                        local_last_row_occurrences.append((self.get_row(row),self.get_col(col)))
                        self.matrix[self.get_row(row)][self.get_col(col)].type = "Last Row"
                        self.matrix[self.get_row(row)][self.get_col(col)].value = last_row_value
        # print(f"local_column_occurrences: {local_column_occurrences} \n")
        # print(f"local_row_name_occurrences: {local_row_name_occurrences} \n")
        # print(f"local_last_row_occurrences: {local_last_row_occurrences} \n")
        return local_column_occurrences, local_row_name_occurrences, local_last_row_occurrences
    
    def find_anchor_point(self, local_column_occurrences):
        """ Finds the top-left most column header and uses that as an anchor-point."""
        anchor_point = min(local_column_occurrences, key=lambda x: (x[0], x[1]))
        return anchor_point

    def find_column(self, local_column_occurrences, local_column_names):
        """ Finds the columns for a given table."""
        
        # Overall method: 
        # For each coordinate in local column occurrences
            # Create a new Counter object for each possible column sequence that counts # of each type of column name
            # Create a queue to store coordinates to be processed
            # Start a BFS from coordinate and check the 3x3 grid around the coordinate for the next column header
                # If the next column exists, add it to the column queue and decrement the corresponding key-value in Counter by 1
                # Continue until queue is empty or Counter reaches 0 completely
            # Compare column sequence found vs number of total column names
            # Track the most accurate column sequence found
        best_sequence = deque()
        best_sequence_score = inf
        column_coordinates_seen = []
        for coordinate in local_column_occurrences:
            if coordinate in column_coordinates_seen or coordinate in self.global_columns_found:
                continue
            else:
                #BFS setup
                column_header_counter = Counter(local_column_names)
                column_header_list = []
                # column_header_queue = deque()
                # column_header_queue.append(coordinate)
                # final_sequence = deque()
                # #BFS
                # while len(column_header_queue) > 0:
                #     if len(final_sequence) == len(local_column_names):
                #         break
                #     start_coord = column_header_queue.popleft()
                #     final_sequence.append(start_coord)
                #     for row in range(start_coord[0] - 1, start_coord[0] + 2):
                #         for col in range(start_coord[1] - 1, start_coord[1] + 2):
                #             if (row, col) == coordinate:
                #                 pass
                #             elif row < 0 or row > self.max_row - 1 or col < 0 or col > self.max_col - 1:
                #                 pass
                #             elif self.matrix[row][col].type == "Column" and column_header_counter[self.matrix[row][col].value] > 0:
                #                 column_header_queue.append((row, col))
                #                 column_header_counter[self.matrix[row][col].value] -= 1
                #     print(f"Column Header Queue: {final_sequence}")
                
                for row in range(coordinate[0] - 1, coordinate[0] + 2):
                    for col in range(0, self.max_col):
                        if self.matrix[row][col].type == "Column" and self.matrix[row][col].value in column_header_counter and column_header_counter[self.matrix[row][col].value] > 0:
                            column_header_list.append((row, col))
                            column_header_counter[self.matrix[row][col].value] -= 1
                # Add sequence to list of coordinates already seen so we don't check through already seen sequences
                column_coordinates_seen.extend(column_header_list)
                
                # Calculate the score for this sequence
                sequence_score = len(column_header_list) / len(local_column_names)
                #If the score is less than the best score, update the best score and the best sequence
                if abs(1 - sequence_score) < best_sequence_score:
                    best_sequence = column_header_list
                    best_sequence_score = abs(1 - sequence_score)
        best_sequence = list(best_sequence)
        self.global_columns_found.extend(best_sequence)

        return best_sequence
        

    
    def form_table(self, filtered_column_occurrences, num_columns):
        """ Filter out near identical tables from the table."""
        table = Table()
        table.column_headers = filtered_column_occurrences
        table.num_columns = num_columns
        return table
    

    def display_table_coordinates(self):
        index = 0
        for table in self.tables:
            all_coords = table.column_headers + table.row_names + table.last_row
            if all_coords:
                
                # Find the top-left most coordinate (minimum row and column)
                top_left = min(table.column_headers, key=lambda x: (x[0], x[1]))
                top_left_xl = self.xy_to_excel(top_left[1] + 1, top_left[0] + 1)
                
                # Find the bottom-left most coordinate (maximum row, minimum column)
                bottom_left = min([(coord[0], coord[1]) for coord in all_coords if coord[0] == max(all_coords, key=lambda x: x[0])[0]], key=lambda x: x[1])
                bottom_left = (bottom_left[0], top_left[1])
                bottom_left_xl = self.xy_to_excel(bottom_left[1] + 1, bottom_left[0] + 1)
                
                # Find the top-right most coordinate (minimum row, maximum column)
                top_right_row = top_left[0]
                top_right_col = top_left[1] + table.num_columns - 1
                top_right_xl = self.xy_to_excel(top_right_col + 1, top_right_row + 1)

                bottom_right_xl = self.xy_to_excel(top_right_col + 1, bottom_left[0] + 1)
                
                self.corner_coordinates.append((self.ws[top_left_xl].value, self.ws[bottom_left_xl].value, self.ws[top_right_xl].value, self.ws[bottom_right_xl].value))
                print(f"Table {index}:")    
                print(f"Top Left Coordinate: {top_left_xl}, Actual Value: {self.ws[top_left_xl].value}")
                print(f"Bottom Left Coordinate: {bottom_left_xl}, Actual Value: {self.ws[bottom_left_xl].value}")
                print(f"Top Right Coordinate: {top_right_xl}, Actual Value: {self.ws[top_right_xl].value}")
                print(f"Bottom Right Coordinate: {bottom_right_xl}, Actual Value: {self.ws[bottom_right_xl].value}")
                print("\n")
                index += 1

    def process_tables(self):
        self.read_json()
        self.populate_cells()
        # global_columns = []
        # for _, table_data in self.json_data.items():
        #     columns, row_names, last_row = self.get_table_data(table_data)
        #     global_columns.extend(columns)
        for table_name, table_data in self.json_data.items():
            print(f"Table Name: {table_name}")
            columns, row_names, last_row = self.get_table_data(table_data)
            num_columns = len(columns)
            print(f"With columns of {columns}")
            print(f"With row names of {row_names}")
            print(f"With last rows of {last_row}")
            local_column_occurrences, local_row_name_occurrences, local_last_row_occurrences = self.pre_pass(columns, row_names, last_row)
            # self.print_matrix()
            print(f"Local Column Occurrences: {local_column_occurrences} \n")
            print(f"Local Row Name Occurrences: {local_row_name_occurrences} \n")
            print(f"Local Last Row Occurrences: {local_last_row_occurrences} \n")
            column_sequence = self.find_column(local_column_occurrences, columns)
            print(f"Columns Found: {column_sequence} \n")
            filtered_table = self.form_table(column_sequence, num_columns)
            self.tables.append(filtered_table)
        #self.display_table_coordinates()
        

def main():
    ex = ExcelProcessor(
        xlsx_file=argv[1],
        json_file=argv[2]
    )
    ex.process_tables()

if __name__ == "__main__":
    main()